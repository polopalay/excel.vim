import sys
import warnings
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def get_sheet(wb, sheet_name=None) -> Worksheet:
    if sheet_name:
        try:
            return wb[sheet_name]
        except KeyError:
            raise RuntimeError(f"Sheet not found: {sheet_name}")

    return wb.worksheets[0]


# Tắt warning của openpyxl về Data Validation extension không được hỗ trợ
# (warning này không ảnh hưởng đến việc đọc/ghi dữ liệu, chỉ làm nhiễu output)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def ensure_workbook(path, sheet_name=None):
    """
    Đảm bảo file .xlsx tồn tại trước khi xử lý.
    Nếu file chưa tồn tại (ví dụ user tạo file mới trong Vim), tạo 1 workbook
    rỗng với 1 sheet và 1 cell A1 trống, rồi lưu lại.
    """
    if os.path.exists(path):
        return
    wb = Workbook()
    ws = get_sheet(wb, sheet_name)
    ws["A1"] = ""
    wb.save(path)


def get_merge_ranges(ws):
    """
    Lấy danh sách tất cả các vùng merge cell (gộp ô) trong worksheet.
    Dùng cách "chuẩn" qua openpyxl (ws.merged_cells.ranges) -> chậm hơn
    fast_get_merge_ranges() vì phải load toàn bộ workbook.

    Trả về: list các dict {min_row, max_row, min_col, max_col}
    """
    ranges = []
    for mc in ws.merged_cells.ranges:
        ranges.append(
            {
                "min_row": mc.min_row,
                "max_row": mc.max_row,
                "min_col": mc.min_col,
                "max_col": mc.max_col,
            }
        )
    return ranges


def build_display_plan(ranges, max_row, max_col):
    """
    Xây dựng "kế hoạch hiển thị" cho worksheet dựa trên các vùng merge.

    Khi 1 vùng ô được merge (gộp), chỉ ô đầu tiên (top-left) của vùng đó
    mới thực sự chứa giá trị; các ô còn lại trong vùng merge là "ảo"
    (không hiển thị nội dung). Hàm này tính trước những ô nào cần ẩn
    nội dung, và những cột nào cần ẩn hẳn (không hiển thị cả cột) khi
    đó là 1 vùng merge ngang (horizontal merge, gộp nhiều cột trên 1 dòng).

    Trả về:
        skip:
            Set các toạ độ (row, col) thuộc vùng merge nhưng KHÔNG phải
            ô top-left -> các ô này sẽ hiển thị giá trị rỗng.

        top_left:
            Dict map (min_row, min_col) -> thông tin vùng merge tương ứng.
            (Hiện tại chưa được dùng trực tiếp ở nơi gọi nhưng vẫn trả về
            để dùng khi cần mở rộng logic sau này.)

        hskip_cols_by_row:
            Dict map số dòng -> set các cột cần ẩn HOÀN TOÀN trên dòng đó,
            vì các cột này bị 1 vùng merge ngang "nuốt" vào ô đầu tiên.
            Ví dụ: merge A1:C1 (ngang) -> dòng 1 sẽ ẩn cột B, C khi hiển thị,
            chỉ còn lại 1 ô gộp đại diện ở cột A.
    """
    skip = set()
    hskip_cols_by_row = {}
    top_left = {}
    for rg in ranges:
        top_left[(rg["min_row"], rg["min_col"])] = rg
        # Vùng merge được coi là "ngang" nếu nó trải rộng qua nhiều cột
        # (min_col khác max_col), bất kể có trải qua nhiều dòng hay không
        is_horizontal = rg["min_col"] != rg["max_col"]
        for r in range(rg["min_row"], rg["max_row"] + 1):
            for c in range(rg["min_col"], rg["max_col"] + 1):
                # Mọi ô trong vùng merge, trừ ô top-left, đều bị đánh dấu "skip"
                if (r, c) != (rg["min_row"], rg["min_col"]):
                    skip.add((r, c))
                # Nếu là merge ngang, các cột (trừ cột đầu) sẽ bị ẩn hẳn
                # trên TỪNG dòng thuộc vùng merge này
                if is_horizontal and c != rg["min_col"]:
                    hskip_cols_by_row.setdefault(r, set()).add(c)
    return skip, top_left, hskip_cols_by_row


def fast_get_merge_ranges(path):
    """
    Đọc trực tiếp các vùng merge từ cấu trúc XML bên trong file XLSX,
    thay vì load toàn bộ workbook bằng openpyxl (cách này nhanh hơn
    nhiều vì file .xlsx về bản chất là 1 file zip chứa các file XML,
    và ta chỉ cần đọc đúng phần worksheet XML cần thiết).

    Quy trình:
        1. Mở file .xlsx như 1 file zip.
        2. Đọc xl/workbook.xml để biết danh sách các sheet và sheet
           nào đang active (đang được chọn khi file được lưu lần cuối).
        3. Đọc xl/_rels/workbook.xml.rels để tìm đường dẫn file XML
           thực tế tương ứng với sheet active đó (vì workbook.xml chỉ
           lưu r:id tham chiếu, không lưu trực tiếp đường dẫn).
        4. Đọc file worksheet XML đó, dùng regex để tìm tất cả thẻ
           <mergeCell ref="A1:C1"/> rồi quy đổi sang toạ độ số (row, col).

    Trả về:
        List các vùng merge nếu thành công.

        None nếu gặp bất kỳ điều gì bất thường (thiếu file, sai cấu trúc,
        XML lỗi, v.v.) -> để hàm gọi nó tự fallback sang cách chuẩn
        bằng openpyxl (get_merge_ranges) cho an toàn.
    """
    import zipfile
    import re
    from xml.etree import ElementTree as ET

    try:
        with zipfile.ZipFile(path) as z:
            names = set(z.namelist())
            if "xl/workbook.xml" not in names:
                return None

            wb_xml = z.read("xl/workbook.xml").decode("utf-8")
            wb_root = ET.fromstring(wb_xml)
            # Namespace chính của spreadsheet XML (theo chuẩn OpenXML)
            ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            # Namespace dùng cho thuộc tính r:id (liên kết tới file quan hệ .rels)
            ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

            sheets = wb_root.findall("m:sheets/m:sheet", ns)
            if not sheets:
                return None

            # Xác định sheet đang active. Nếu không có thông tin activeTab
            # (book_views), mặc định lấy sheet đầu tiên (index 0)
            active_idx = 0
            book_views = wb_root.find("m:bookViews/m:workbookView", ns)
            if book_views is not None and book_views.get("activeTab") is not None:
                try:
                    active_idx = int(book_views.get("activeTab"))
                except ValueError:
                    active_idx = 0
            if active_idx >= len(sheets):
                active_idx = 0

            # Lấy r:id của sheet active -> dùng để tra trong file .rels
            r_id = sheets[active_idx].get(f"{{{ns_r}}}id")
            if not r_id:
                return None

            rels_path = "xl/_rels/workbook.xml.rels"
            if rels_path not in names:
                return None
            rels_xml = z.read(rels_path).decode("utf-8")
            rels_root = ET.fromstring(rels_xml)
            target = None
            for rel in rels_root:
                if rel.get("Id") == r_id:
                    target = rel.get("Target")
                    break
            if not target:
                return None

            # Chuẩn hoá đường dẫn target: bỏ dấu '/' ở đầu (nếu có) và
            # đảm bảo có tiền tố "xl/" (vì Target trong .rels thường là
            # đường dẫn tương đối tính từ thư mục xl/)
            target = target.lstrip("/")
            if not target.startswith("xl/"):
                target = "xl/" + target
            if target not in names:
                return None

            sheet_xml = z.read(target).decode("utf-8")

        # Thẻ mergeCell thường nằm gần cuối file worksheet XML, nhưng để
        # an toàn (tránh phụ thuộc vào thứ tự cụ thể), regex quét toàn bộ
        # nội dung file để tìm tất cả các thẻ <mergeCell ref="...">
        refs = re.findall(r'<mergeCell[^>]*\sref="([A-Z]+\d+):([A-Z]+\d+)"', sheet_xml)

        def col_to_num(col_str):
            """Quy đổi tên cột dạng chữ (A, B, ..., Z, AA, AB, ...) sang số thứ tự cột."""
            num = 0
            for ch in col_str:
                num = num * 26 + (ord(ch) - ord("A") + 1)
            return num

        ranges = []
        for start, end in refs:
            # Tách phần chữ (tên cột) và phần số (số dòng) từ mỗi ô,
            # ví dụ "B12" -> nhóm chữ "B", nhóm số "12"
            sm = re.match(r"([A-Z]+)(\d+)", start)
            em = re.match(r"([A-Z]+)(\d+)", end)
            if not sm or not em:
                return None
            ranges.append(
                {
                    "min_col": col_to_num(sm.group(1)),
                    "min_row": int(sm.group(2)),
                    "max_col": col_to_num(em.group(1)),
                    "max_row": int(em.group(2)),
                }
            )
        return ranges
    except Exception:
        # Bất kỳ lỗi nào xảy ra (file hỏng, XML sai cấu trúc, thiếu key, ...)
        # đều coi như "không đọc được bằng cách nhanh" -> trả None để fallback
        return None


def render(rows):
    """
    Chuyển 1 danh sách các dòng (mỗi dòng là list các string) thành
    1 chuỗi văn bản dạng bảng ASCII, có viền +---+---+ và dấu | phân cách cột.

    Mỗi cột sẽ tự động được căn độ rộng theo giá trị dài nhất trong cột đó
    trên toàn bộ các dòng (left-justify, căn trái).
    """
    if not rows:
        return ""

    # Tính độ rộng tối đa cần thiết cho mỗi cột (dựa theo dòng có nhiều cột nhất)
    widths = [0] * max(len(r) for r in rows)
    for row in rows:
        for i, v in enumerate(row):
            widths[i] = max(widths[i], len(v))

    # Dòng viền ngang, ví dụ: +-----+-------+---+
    border = "+" + "+".join("-" * (w + 2) for w in widths) + "+"

    out = [border]
    for row in rows:
        parts = ["|"]
        for i, v in enumerate(row):
            parts.append(" ")
            parts.append(v.ljust(widths[i]))  # căn trái theo độ rộng cột
            parts.append(" |")
        out.append("".join(parts))
        out.append(
            border
        )  # vẽ viền ngang sau MỖI dòng dữ liệu (không chỉ đầu/cuối bảng)
    return "\n".join(out)


def open_xlsx(path, sheet_name=None):
    ensure_workbook(path)

    wb = load_workbook(path, data_only=True, keep_links=False)
    ws = get_sheet(wb, sheet_name)

    if sheet_name:
        ranges = get_merge_ranges(ws)
    else:
        ranges = fast_get_merge_ranges(path)

        if ranges is None:
            ranges = get_merge_ranges(ws)

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    skip, _, hskip_cols_by_row = build_display_plan(ranges, max_row, max_col)

    display_rows = []

    # Lặp qua từng dòng của worksheet để xây dựng nội dung hiển thị
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row), start=1):
        cols_to_drop = hskip_cols_by_row.get(r)
        display_row = []
        row_len = len(row)
        for c in range(1, max_col + 1):
            # Cột này bị merge ngang "nuốt" trên dòng hiện tại -> bỏ qua hẳn,
            # không thêm vào display_row (khác với "skip" - vẫn thêm nhưng để rỗng)
            if cols_to_drop and c in cols_to_drop:
                continue
            if (r, c) in skip:
                # Ô thuộc vùng merge nhưng không phải ô top-left -> hiển thị rỗng
                display_row.append("")
            else:
                v = row[c - 1].value if c <= row_len else None
                display_row.append(
                    ""
                    if v is None
                    else str(v).replace("\r", " ").replace("\n", " \\n ")
                    # Thay \r bằng khoảng trắng và \n bằng " \n " (dạng escape
                    # hiển thị được trên 1 dòng) để giữ mỗi cell trong bảng ASCII
                    # luôn nằm trên đúng 1 dòng text, không bị vỡ layout bảng
                )
        display_rows.append(display_row)

    wb.close()
    if not display_rows:
        display_rows = [[""]]

    print(render(display_rows))


def parse_ascii(lines):
    """
    Phân tích ngược lại: từ các dòng text dạng bảng ASCII (do render() tạo ra,
    đã được user chỉnh sửa trong Vim) -> trích xuất lại thành list các dòng
    dữ liệu (mỗi dòng là list giá trị string, đã strip khoảng trắng).

    Chỉ những dòng bắt đầu bằng '|' (dòng dữ liệu) mới được xử lý;
    các dòng viền (bắt đầu bằng '+') sẽ bị bỏ qua.
    """
    rows = []
    for line in lines:
        line = line.rstrip()
        if not line.startswith("|"):
            continue
        # Cắt theo dấu '|', bỏ phần tử đầu và cuối (rỗng, do dòng bắt đầu/kết
        # thúc bằng '|'), chỉ giữ lại các giá trị cell ở giữa
        cols = line.split("|")[1:-1]
        rows.append([c.strip() for c in cols])
    return rows


def compute_mergeinfo(path, sheet_name=None):
    """
    Tính toán thông tin merge và bản đồ ánh xạ cột (column mapping) trực tiếp
    từ file .xlsx hiện có, dùng cùng logic với open_xlsx(). Không cache ra
    đĩa -> không sinh ra file tạm nào.

    col_map_per_row giúp save_xlsx() biết: ứng với mỗi cột HIỂN THỊ trong
    bảng ASCII (sau khi đã ẩn các cột bị merge ngang nuốt), nó tương ứng
    với cột THỰC TẾ nào trong file Excel gốc.
    """
    if sheet_name:
        wb = load_workbook(path, data_only=True, keep_links=False)
        ws = get_sheet(wb, sheet_name)
        ranges = get_merge_ranges(ws)
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        wb.close()
    else:
        ranges = fast_get_merge_ranges(path)
        if ranges is None:
            wb = load_workbook(path, data_only=True, keep_links=False)
            ws = get_sheet(wb)
            ranges = get_merge_ranges(ws)
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            wb.close()
        else:
            wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
            ws = get_sheet(wb)
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            wb.close()

    _, _, hskip_cols_by_row = build_display_plan(ranges, max_row, max_col)

    # Với mỗi dòng, xây danh sách các cột "thực" còn được hiển thị
    # (loại bỏ các cột bị ẩn do nằm trong vùng merge ngang)
    col_map_per_row = []
    for r in range(1, max_row + 1):
        cols_to_drop = hskip_cols_by_row.get(r, set())
        col_map = [c for c in range(1, max_col + 1) if c not in cols_to_drop]
        col_map_per_row.append(col_map)

    return {
        "max_row": max_row,
        "max_col": max_col,
        "ranges": ranges,
        "col_map_per_row": col_map_per_row,
    }


def save_xlsx(xlsx_file, txt_file, sheet_name=None):
    """
    Hàm chính xử lý lệnh 'save': đọc nội dung bảng ASCII đã được user chỉnh
    sửa trong Vim (txt_file), rồi ghi ngược lại các giá trị vào file .xlsx
    gốc (xlsx_file), đồng thời khôi phục lại đúng các vùng merge ban đầu.

    Quy trình:
        1. Đọc file text (txt_file) -> parse thành các dòng dữ liệu thô.
        2. Lấy thông tin merge/col_map TỪ FILE XLSX GỐC (trước khi sửa)
           để biết cấu trúc merge hiện tại ra sao.
        3. Mở workbook, gỡ HẾT các merge hiện có (để có thể ghi giá trị
           tự do vào từng cell mà không bị chặn bởi merge cũ).
        4. Ghi giá trị mới vào đúng cell thực tế, dựa theo col_map_per_row
           (ánh xạ từ vị trí cột hiển thị -> vị trí cột thực trong Excel).
        5. Áp lại các vùng merge gốc (dùng lại đúng ranges đã lấy ở bước 2).
        6. Lưu file.
    """
    ensure_workbook(xlsx_file)
    with open(txt_file, encoding="utf-8") as f:
        lines = f.readlines()
    new_rows = parse_ascii(lines)

    # Lấy thông tin merge/col_map dựa trên cấu trúc HIỆN TẠI của file gốc
    # (trước khi bị ghi đè), để biết cách ánh xạ cột hiển thị -> cột thực
    info = compute_mergeinfo(
        xlsx_file,
        sheet_name,
    )

    wb = load_workbook(xlsx_file)
    ws = get_sheet(wb, sheet_name)

    # Gỡ toàn bộ merge hiện có trước khi ghi giá trị mới, vì openpyxl
    # không cho ghi trực tiếp vào các ô "ảo" (không phải top-left) của
    # 1 vùng đang merge
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))

    # Nếu số dòng trong col_map_per_row (tính từ file gốc) đủ để khớp với
    # số dòng mới (new_rows) thì dùng ánh xạ đó. Nếu không (ví dụ user đã
    # thêm/xoá dòng theo cách thủ công ngoài các hàm insert_row), dùng
    # fallback ánh xạ 1:1 (cột hiển thị thứ i = cột thực thứ i) để tránh lỗi.
    if len(info["col_map_per_row"]) >= len(new_rows):
        col_map_per_row = info["col_map_per_row"]
    else:
        # fallback: số dòng không khớp -> map 1:1 như cũ
        col_map_per_row = [list(range(1, len(row) + 1)) for row in new_rows]

    # đồng bộ số dòng
    old_row_count = info["max_row"]
    new_row_count = len(new_rows)

    if new_row_count < old_row_count:
        ws.delete_rows(new_row_count + 1, old_row_count - new_row_count)

    elif new_row_count > old_row_count:
        ws.insert_rows(old_row_count + 1, new_row_count - old_row_count)

    # đồng bộ số cột
    old_col_count = info["max_col"]
    new_col_count = max((len(r) for r in new_rows), default=0)

    if new_col_count < old_col_count:
        ws.delete_cols(new_col_count + 1, old_col_count - new_col_count)

    elif new_col_count > old_col_count:
        ws.insert_cols(old_col_count + 1, new_col_count - old_col_count)

    # ghi dữ liệu
    for r_idx, row in enumerate(new_rows):
        col_map = (
            col_map_per_row[r_idx]
            if r_idx < len(col_map_per_row)
            else list(range(1, len(row) + 1))
        )

        for i, value in enumerate(row):
            real_col = col_map[i] if i < len(col_map) else i + 1
            value = value.replace("\\n", "\n")
            cell = ws.cell(r_idx + 1, real_col)
            cell.value = value

    # Áp lại các vùng merge gốc (đã gỡ ở bước trên) để giữ đúng layout ban đầu
    for rg in info["ranges"]:
        try:
            start = f"{get_column_letter(rg['min_col'])}{rg['min_row']}"
            end = f"{get_column_letter(rg['max_col'])}{rg['max_row']}"
            ws.merge_cells(f"{start}:{end}")
        except Exception:
            # Nếu 1 vùng merge nào đó không áp lại được (ví dụ do dữ liệu
            # đã thay đổi cấu trúc dòng/cột), bỏ qua để không làm hỏng
            # toàn bộ quá trình lưu
            pass

    wb.save(xlsx_file)


def shift_merges_for_inserted_row(old_ranges, row_num):
    """
    Tính lại toạ độ các vùng merge sau khi có 1 dòng mới được chèn vào
    vị trí row_num.

    Quy tắc:
        - Nếu vùng merge nằm HOÀN TOÀN sau vị trí chèn (min_row >= row_num)
          -> toàn bộ vùng bị đẩy xuống 1 dòng (cả min_row và max_row +1).
        - Nếu vùng merge "bao trùm" vị trí chèn (min_row < row_num nhưng
          max_row >= row_num) -> dòng mới được chèn vào GIỮA vùng merge đó,
          nên vùng merge được MỞ RỘNG thêm 1 dòng (chỉ max_row +1, giữ
          nguyên min_row) để dòng mới cũng thuộc vùng merge này.
        - Nếu vùng merge nằm hoàn toàn trước vị trí chèn (max_row < row_num)
          -> không thay đổi.
    """
    new_ranges = []
    for rg in old_ranges:
        min_row, max_row = rg["min_row"], rg["max_row"]
        if min_row >= row_num:
            min_row += 1
            max_row += 1
        elif max_row >= row_num:
            # vùng merge bao trùm vị trí chèn -> mở rộng thêm 1 dòng
            max_row += 1
        new_ranges.append(
            {
                "min_row": min_row,
                "max_row": max_row,
                "min_col": rg["min_col"],
                "max_col": rg["max_col"],
            }
        )
    return new_ranges


def shift_merges_for_inserted_col(old_ranges, col_num):
    """
    Tương tự shift_merges_for_inserted_row(), nhưng áp dụng cho việc
    chèn thêm 1 CỘT mới vào vị trí col_num (dịch theo chiều cột thay
    vì theo chiều dòng).
    """
    new_ranges = []
    for rg in old_ranges:
        min_col, max_col = rg["min_col"], rg["max_col"]
        if min_col >= col_num:
            min_col += 1
            max_col += 1
        elif max_col >= col_num:
            max_col += 1
        new_ranges.append(
            {
                "min_row": rg["min_row"],
                "max_row": rg["max_row"],
                "min_col": min_col,
                "max_col": max_col,
            }
        )
    return new_ranges


def list_sheets(path):
    wb = load_workbook(path, read_only=True)

    for ws in wb.worksheets:
        print(ws.title)

    wb.close()


# ----------------------------------------------------------------------------
# Entry point: nhận lệnh từ dòng lệnh (do excelPlugin.vim gọi qua systemlist())
#
# Cách dùng:
#   python excel.py open <xlsx_file> [sheet_name]: mở file .xlsx và hiển thị nội dung sheet dưới dạng bảng ASCII trong Vim
#   python excel.py save <xlsx_file> <txt_file> [sheet_name]: lưu nội dung đã chỉnh sửa trong bảng ASCII (txt_file) trở lại file .xlsx, giữ nguyên cấu trúc merge ban đầu
#   python excel.py sheets <xlsx_file>: liệt kê tên tất cả các sheet trong file .xlsx (dùng để user biết đặt sheet_name thế nào khi mở)
# ----------------------------------------------------------------------------
if __name__ == "__main__":
    mode = sys.argv[1]
    if mode == "open":
        open_xlsx(
            sys.argv[2],
            sys.argv[3] if len(sys.argv) > 3 else None,
        )

    elif mode == "save":
        save_xlsx(
            sys.argv[2],
            sys.argv[3],
            sys.argv[4] if len(sys.argv) > 4 else None,
        )

    elif mode == "sheets":
        list_sheets(sys.argv[2])
